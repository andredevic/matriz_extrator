#!/usr/bin/env python3
"""Extrai dados de Matrizes Energéticas (Excel) e gera planilha consolidada.

Lê: .xls (converte), .xlsx, .xlsm
Entrada: ./planilhas/
Saída: ./saida/matriz_consolidada.xlsx
Convertidos: ./convertidos/

IMPORTANTE:
- Conversão .xls usa Excel via COM (pywin32). Alguns arquivos podem travar por prompts.
- Este script usa TIMEOUT e mata o Excel do arquivo travado, seguindo para o próximo.

Mapeamento (linhas a partir da 11):
- Tag Equipamento: B (fill down)
- Descrição Equipamento: C + D (fill down)
- Tag Fonte: E + F + G
- Descrição Fonte: H + I + J + K + L
- Como Bloquear: M + N
- Onde Bloquear/TAG: O + P + Q + R + S
- Tipo de Bloqueio: T + U
- Como Desbloquear: Z + AA
"""

from __future__ import annotations

import re
import time
import subprocess
from pathlib import Path
from typing import Iterable, Optional, List, Tuple

import pandas as pd
from openpyxl import load_workbook


# =========================
# Configurações
# =========================
BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "planilhas"
OUTPUT_DIR = BASE_DIR / "saida"
OUTPUT_FILE = OUTPUT_DIR / "matriz_consolidada.xlsx"
CONVERTED_DIR = BASE_DIR / "convertidos"

START_ROW = 11

# Timeout por arquivo .xls durante a conversão (segundos)
XLS_CONVERT_TIMEOUT_SEC = 40

COLS_EQUIP_TAG = ["B"]
COLS_EQUIP_DESC = ["C", "D"]

COLS_FONTE_TAG = ["E", "F", "G"]
COLS_FONTE_DESC = ["H", "I", "J", "K", "L"]

COLS_COMO_BLOQUEAR = ["M", "N"]
COLS_ONDE_BLOQUEAR = ["O", "P", "Q", "R", "S"]
COLS_TIPO_BLOQUEIO = ["T", "U"]
COLS_COMO_DESBLOQUEAR = ["Z", "AA"]

ALL_RELEVANT_COLS = sorted(
    set(
        COLS_EQUIP_TAG
        + COLS_EQUIP_DESC
        + COLS_FONTE_TAG
        + COLS_FONTE_DESC
        + COLS_COMO_BLOQUEAR
        + COLS_ONDE_BLOQUEAR
        + COLS_TIPO_BLOQUEIO
        + COLS_COMO_DESBLOQUEAR
    )
)

EMPTY_TOKENS = {"-", "—", "–", "n/a", "na", "null", "none"}

FOOTER_KEYWORDS = [
    "LEGENDA",
    "ELABORADOR",
    "REVISOR",
    "APROVADOR",
    "PROVIDENCIAS",
    "PROVIDÊNCIAS",
    "DISPOSITIVO",
    "LAYOUT",
    "PÁGINA",
    "PAGINA",
]


# =========================
# Utilitários
# =========================
def normalize_cell(value) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        s = str(value)
    else:
        s = str(value)

    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    if s.lower() in EMPTY_TOKENS:
        return None
    return s


def join_valid(values: Iterable[Optional[str]], sep: str = " ") -> Optional[str]:
    parts: List[str] = []
    for v in values:
        nv = normalize_cell(v)
        if nv is not None:
            parts.append(nv)
    if not parts:
        return None
    out = sep.join(parts)
    out = re.sub(r"\s+", " ", out).strip()
    return out or None


def get_group(ws, row: int, cols: list[str], sep: str = " ") -> Optional[str]:
    vals = [ws[f"{c}{row}"].value for c in cols]
    return join_valid(vals, sep=sep)


def row_has_any_data(ws, row: int) -> bool:
    for c in ALL_RELEVANT_COLS:
        v = normalize_cell(ws[f"{c}{row}"].value)
        if v is not None:
            return True
    return False


def row_has_footer_marker(ws, row: int) -> bool:
    check_cols = ["A", "B", "C", "D", "E", "H", "M", "O", "T", "Z", "AA"]
    texts: List[str] = []
    for c in check_cols:
        v = normalize_cell(ws[f"{c}{row}"].value)
        if v:
            texts.append(v.upper())
    if not texts:
        return False
    joined = " | ".join(texts)
    return any(k in joined for k in FOOTER_KEYWORDS)


def list_input_files() -> List[Path]:
    if not INPUT_DIR.exists():
        return []
    files: List[Path] = []
    for pattern in ("*.xls", "*.xlsx", "*.xlsm"):
        for p in INPUT_DIR.rglob(pattern):
            if p.name.startswith("~$"):
                continue
            files.append(p)
    return sorted(set(files))


def kill_excel_processes() -> None:
    """Mata EXCEL.EXE para destravar conversões presas."""
    # /F força, /IM filtra por nome do executável
    subprocess.run(
        ["taskkill", "/F", "/IM", "EXCEL.EXE"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        shell=False,
    )


# =========================
# Conversão .xls -> .xlsx (com timeout)
# =========================
def convert_xls_to_xlsx(xls_path: Path) -> Path:
    CONVERTED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = CONVERTED_DIR / (xls_path.stem + ".xlsx")

    # Se já existe convertido, reaproveita
    if out_path.exists():
        return out_path

    try:
        import win32com.client  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "Para ler .xls, instale pywin32: python -m pip install pywin32"
        ) from e

    # Garantia: não deixar excel antigo aberto travando
    # (opcional, mas ajuda em lote)
    # kill_excel_processes()

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    started = time.time()
    wb = None

    try:
        # Open(..., UpdateLinks=0) evita prompts de links
        wb = excel.Workbooks.Open(str(xls_path), UpdateLinks=0, ReadOnly=True)

        # Timeout manual (COM pode travar sem levantar exceção)
        while True:
            if time.time() - started > XLS_CONVERT_TIMEOUT_SEC:
                raise TimeoutError(f"Timeout convertendo .xls (> {XLS_CONVERT_TIMEOUT_SEC}s)")
            # Se chegou aqui, workbook abriu — sai do loop
            break

        # 51 => .xlsx
        wb.SaveAs(str(out_path), FileFormat=51)
        wb.Close(SaveChanges=False)
        wb = None

    except TimeoutError:
        # Mata excel travado
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        kill_excel_processes()
        raise

    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass

    return out_path


# =========================
# Processamento
# =========================
def process_workbook(path_xlsx: Path, source_file_name: str) -> list[dict]:
    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.worksheets[0]

    records: list[dict] = []
    last_equip_tag: Optional[str] = None
    last_equip_desc: Optional[str] = None

    max_row = ws.max_row or START_ROW

    for r in range(START_ROW, max_row + 1):
        if row_has_footer_marker(ws, r):
            break

        if not row_has_any_data(ws, r):
            continue

        equip_tag = get_group(ws, r, COLS_EQUIP_TAG)
        equip_desc = get_group(ws, r, COLS_EQUIP_DESC)

        if equip_tag is None:
            equip_tag = last_equip_tag
        else:
            last_equip_tag = equip_tag

        if equip_desc is None:
            equip_desc = last_equip_desc
        else:
            last_equip_desc = equip_desc

        fonte_tag = get_group(ws, r, COLS_FONTE_TAG)  # vazio => None (null)
        fonte_desc = get_group(ws, r, COLS_FONTE_DESC)
        como_bloquear = get_group(ws, r, COLS_COMO_BLOQUEAR)
        onde_bloquear = get_group(ws, r, COLS_ONDE_BLOQUEAR)
        tipo_bloqueio = get_group(ws, r, COLS_TIPO_BLOQUEIO)
        como_desbloquear = get_group(ws, r, COLS_COMO_DESBLOQUEAR)

        record = {
            "Arquivo de Origem": source_file_name,
            "Tag do Equipamento": equip_tag,
            "Descrição do Equipamento": equip_desc,
            "Tag da Fonte de Energia": fonte_tag,
            "Descrição da Fonte de Energia": fonte_desc,
            "Como Bloquear": como_bloquear,
            "Onde Bloquear / TAG": onde_bloquear,
            "Tipo de Bloqueio": tipo_bloqueio,
            "Como Desbloquear": como_desbloquear,
        }

        has_source_info = any(
            record[k] is not None
            for k in [
                "Tag da Fonte de Energia",
                "Descrição da Fonte de Energia",
                "Como Bloquear",
                "Onde Bloquear / TAG",
                "Tipo de Bloqueio",
                "Como Desbloquear",
            ]
        )
        if not has_source_info:
            continue

        records.append(record)

    return records


# =========================
# Main
# =========================
def main() -> None:
    if not INPUT_DIR.exists():
        raise SystemExit(f"Pasta de entrada não existe: {INPUT_DIR}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CONVERTED_DIR.mkdir(parents=True, exist_ok=True)

    files = list_input_files()
    if not files:
        raise SystemExit(f"Nenhum .xls/.xlsx/.xlsm encontrado em: {INPUT_DIR}")

    all_records: list[dict] = []
    processed = 0
    converted = 0
    failed: List[Tuple[str, str]] = []

    total = len(files)
    print(f"Encontrados {total} arquivos. Iniciando processamento...")

    for i, f in enumerate(files, start=1):
        print(f"[{i}/{total}] Processando: {f.name}")

        try:
            if f.suffix.lower() == ".xls":
                xlsx_path = convert_xls_to_xlsx(f)
                converted += 1
                recs = process_workbook(xlsx_path, source_file_name=f.name)
            else:
                recs = process_workbook(f, source_file_name=f.name)

            all_records.extend(recs)
            processed += 1

        except Exception as e:
            failed.append((f.name, str(e)))
            print(f"   -> ERRO: {e}")

    df = pd.DataFrame(
        all_records,
        columns=[
            "Arquivo de Origem",
            "Tag do Equipamento",
            "Descrição do Equipamento",
            "Tag da Fonte de Energia",
            "Descrição da Fonte de Energia",
            "Como Bloquear",
            "Onde Bloquear / TAG",
            "Tipo de Bloqueio",
            "Como Desbloquear",
        ],
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consolidado")
        if failed:
            df_err = pd.DataFrame(failed, columns=["Arquivo", "Erro"])
            df_err.to_excel(writer, index=False, sheet_name="Erros")

    print("\nFINALIZADO")
    print(f"OK! Gerado: {OUTPUT_FILE}")
    print(f"Arquivos encontrados: {total} | Processados: {processed} | Convertidos: {converted} | Falharam: {len(failed)}")
    print(f"Linhas consolidadas: {len(df)}")
    if failed:
        print("Obs: veja a aba 'Erros' para detalhes.")


if __name__ == "__main__":
    main()
